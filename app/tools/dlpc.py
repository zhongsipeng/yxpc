'''
拿到游戏名列表，解析游戏名依次执行搜索
搜索步骤.
    根据游戏名获取RJ编号
    根据RJ编号获取游戏主页
    根据游戏主页获取所需数据
读取Excel文件，根据文件输出收集的数据
'''


import requests
from bs4 import BeautifulSoup
from urllib.parse import unquote
import re
import json
from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from copy import copy
import os
from openpyxl.utils import get_column_letter
yxm_key = "游戏名"
yxbh_key = "游戏编号"
game_info_mapping = {
    "游戏名": "yxm",
    "游戏类型": "yxlx",
    "作品类型": "zplx",
    "分类": "flbq",
    "贩卖日": "fsrq",
    "更新信息": "gxrq",
    "社团名": "st",
    "入正地址": "rzdz",
    "文件容量": "wjdx",
    "贩卖数": "fms",
    "评价": "pf",
    "评价人数": "pfrs",
    "支持的语言": "zcyy",
    "游戏编号": "yxbh"
}
def get_game_no(game_name):
    callback = "jQuery151060678_17505131442"
    search = game_name
    url = f"https://www.dlsite.com/suggest/?callback={callback}&term={search}&site=adult-jp&time=17405112&touch=0&_=1741505135927"
    response = requests.get(url)
    data = response.text
    # response = { "text": r'''{"work":[{"work_name":"\u795f\u308a\u306e\u6708 \uff5e\u51cc\u8fb1\u30db\u30e9\u30fc\u30b2\u30fc\u30e0\uff5e","workno":"RJ01044518","maker_name":"\u3064\u3063\u304d\u30fc\u306e\u304a\u8336\u4f1a","maker_id":"RG58056","work_type":"RPG","intro_s":"\u5c11\u5973\u3092\u52a9\u3051\u308b\u305f\u3081\u306b\u4e0d\u6c17\u5473\u306a\u5c4b\u6577\u306b\u6311\u3080\u96ea\u97f3\u3002\u795e\u69d8\u306e\u529b\u3092\u5931\u3063\u305f\u5f7c\u5973\u306b\u6570\u3005\u306e\u5371\u967a\u306a\u7f60\u304c\u8972\u3044\u304b\u304b\u308b","age_category":3,"is_ana":false}],"maker":[{"workno":"RJ01044518","maker_name":"\u3064\u3063\u304d\u30fc\u306e\u304a\u8336\u4f1a","maker_name_kana":"\u30c4\u30c3\u30ad\u30fc\u30ce\u30aa\u30c1\u30e3\u30ab\u30a4","maker_id":"RG58056","age_category":3,"is_ana":false}],"reqtime":17405112}'''}
    # data = response["text"]
    data = data.replace(callback+"(", "").replace(");", "").encode("utf-8").decode("unicode_escape").replace("\r", "").replace("\n", "")
    data = json.loads(data)
    if len(data["work"])>0 :
        return data["work"][0]["workno"]
    else:
        return ""
def get_game_info(rjno):
    data = {}
    if rjno == "":
        return data
    data["游戏编号"] = rjno
    url = f"https://www.dlsite.com/maniax/work/=/product_id/{rjno}.html/?locale=zh_CN"
    url2 = f"https://www.dlsite.com/maniax/product/info/ajax?product_id={rjno}&cdn_cache_min=1"

    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    work_right = soup.find("div", id='work_right')
    table = work_right.find("table", id ="work_outline")

    data["社团名"] = work_right.select_one(".maker_name a").text
    col = ""
    for item in table.find_all("tr"):
        col = item.find("th").text
        match col:
            case '更新信息':
                data[col] = item.find("td").text.strip()
                data[col] = data[col].split("\n")[0]
            # case '贩卖日'|'作品类型'|'支持的语言'|'文件容量':
                pass
            case '分类':
                data[col] = item.find("td").text.strip()
                data[col] = data[col].replace("\n", " ")
            case '支持的语言':
                data[col] = " ".join([i.text for i in item.find_all("a")]).strip()
            case '作品类型':
                zplx = [i.text for i in item.find_all("a")]
                data[col] = " ".join(zplx[1:])
                match zplx[0]:
                    case '模拟':
                        data["游戏类型"] = 'SLG'
                    case '冒险':
                        data["游戏类型"] = 'ADV'
                    case '角色扮演':
                        data["游戏类型"] = 'RPG'
                    case '动作':
                        data["游戏类型"] = 'ACT'
                    case _:
                        data["游戏类型"] = data[col]
            case _:
                data[col] = item.find("td").text.strip()
            
    response = requests.get(url2)
    game_info = json.loads(response.text)

    data["贩卖数"] = game_info[rjno]["dl_count"]
    data["评价"] = game_info[rjno]["rate_average_2dp"]
    data["评价人数"] = game_info[rjno]["rate_count"]
    data["入正地址"] = url
    return data
def build_game_data(rjno):
    if rjno == "":
        return
    # game_info = {'社团名': 'スタジオドビー', '贩卖日': '2023年04月23日', '更新信息': '2023年07月24日', '插画': 'りーん', '声优': '秋山はるる /           浅木式 /           こやまはる /           分倍河原シホ /           水谷六花 /           天川みるく /           来夢ふらん', '年龄指定': 'R18', '作品类型': '有声音 有音乐 有动画', '游戏类型': 'RPG', '文件形式': '软件', '支持的语言': '日文', '分类': '女主人公 羞辱 自慰 淫乱 羞耻/耻辱 多p/乱交 强行 异种X', '文件容量': '5.1GB', '贩卖数': 30626, '评价': 4.55, '评价人数': 9463}
    game_info = get_game_info(rjno)
    data = {}
    for key in game_info_mapping:
        if key == yxm_key:
            continue
        if key in game_info:
            data[game_info_mapping[key]] = game_info[key]
        else:
            data[game_info_mapping[key]] = ""
    data[game_info_mapping[yxbh_key]] = rjno
    return data

def output(template_path, output_path, new_data):
    wb = load_workbook(template_path)
    ws = wb.active

    last_data_row = ws.max_row
    insert_row = 2

    # 动态获取模板的列数（基于表头行）
    header_row = 1
    column_count = ws.max_column  # 或根据表头行实际列数

    pattern = re.compile(r'(http|https):\/\/.*(RJ\d{0,8}).html')
    # 插入数据并复制格式
    reference_row = last_data_row  # 假设最后一行是数据行
    for row_data in new_data:
        for col in range(1, column_count + 1):
            ref_cell = ws.cell(row=reference_row, column=col)
            value = ""
            if col <= len(row_data):
                value = row_data[col-1]
                
            new_cell = ws.cell(row=insert_row, column=col, value=value)
            
            match = pattern.search(value)
            if match:
                new_cell.hyperlink = Hyperlink(ref=get_column_letter(col)+str(reference_row), target=value, tooltip=match[2])
                new_cell.value = match[2]
            
            # 复制样式
            new_cell.font = copy(ref_cell.font)
            new_cell.border = copy(ref_cell.border)
            new_cell.fill = copy(ref_cell.fill)
            new_cell.alignment = copy(ref_cell.alignment)
            new_cell.number_format = ref_cell.number_format

        ws.row_dimensions[insert_row].height = ws.row_dimensions[reference_row].height
        insert_row += 1
    try:
        os.remove(output_path)
    except:
        pass
    wb.save(output_path)
    return output_path
    print(f"数据已插入到 {output_path}")